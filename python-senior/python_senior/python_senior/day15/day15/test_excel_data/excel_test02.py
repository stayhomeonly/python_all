"""
-------------------------------------------------
   Author:Mr.Liu
   date: 2022/8/14-10:29
   contents:
-------------------------------------------------
"""
import openpyxl


# 讲字典转换成对象
class CaseData:
    def __init__(self, dict_case):
        for i in dict_case.items():
            setattr(self, i[0], i[1])


class ReadExcel:
    @classmethod
    def read_data_all(cls, file_name, sheet_name):
        """
        :param file_name:  需要操作的文件
        :param sheet_name: 需要操作的sheet页
        :return: all_case   返回所有的用例
        """
        workbook = openpyxl.load_workbook(file_name)  # 加载文件
        sheet = workbook[sheet_name]  # 选择需要操作的sheet页
        rows = list(sheet.rows)  # 获取所有的行，转换为列表
        all_case = []  # 定义一个空列表用来存储打包后的对象
        titles = []  # 用来存放表头
        for cell in rows[0]:  # rows[0] 代表第一行的数据
            titles.append(cell.value)  # 将单元格的值添加至列表

        for row in rows[1:]:  # rows[1:] 代表除了第一行的其他数据
            data = []  # 定义一个空列表，用来存储数据行的值
            for v in row:  # 循环当前行中的单元格
                data.append(v.value)  # 将单元格的值添加至列表
            case_data = dict(zip(titles, data))  # 将标题和数据进行打包，转换成字典
            case = CaseData(case_data)  # 将转换好的字典转换成对象
            all_case.append(case)  # 将转换好的对象添加致列表中，进行返回
        return all_case

    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):
        """
        :param file_name: 需要操作的文件
        :param sheet_name: 需要操作的sheet页
        :param begin_row: 开始行
        :param end_row: 结束行
        :return: 返回所有的用例
        """
        workbook = openpyxl.load_workbook(file_name)  # 加载文件
        sheet = workbook[sheet_name]  # 选择需要操作的sheet页
        rows = list(sheet.rows)  # 获取所有的行，转换为列表
        all_case = []  # 定义一个空列表用来存储打包后的对象
        titles = []  # 用来存放表头
        for cell in rows[0]:  # rows[0] 代表第一行的数据
            titles.append(cell.value)  # 将单元格的值添加至列表

        for row in rows[begin_row:end_row + 1]:  # 根据开始下标和结束下标进行截取
            data = []  # 定义一个空列表，用来存储数据行的值
            for v in row:  # 循环当前行中的单元格
                data.append(v.value)  # 将单元格的值添加至列表
            case_data = dict(zip(titles, data))  # 将标题和数据进行打包，转换成字典
            case = CaseData(case_data)  # 将转换好的字典转换成对象
            all_case.append(case)  # 将转换好的对象添加致列表中，进行返回
        return all_case

    @classmethod
    def write_data(cls, file_name, sheet_name, row, column, value):
        """
        :param file_name:  需要操作的文件
        :param sheet_name: 需要操作的sheet页
        :param row: 需要被写入的行
        :param column: 需要被写入的列
        :param value: 需要被写入的值
        """
        workbook = openpyxl.load_workbook(file_name)  # 加载文件
        sheet = workbook[sheet_name]  # 选择需要操作的sheet页
        sheet.cell(row=row + 1, column=column, value=value)  # row+1 是为了忽略掉首行，并且和case_id对应
        workbook.save(file_name)


if __name__ == '__main__':
    # res = ReadExcel.read_data_all('./test_case.xlsx', "Sheet1")
    # print(res[2].case_title)
    # res = ReadExcel.read_data_pl('./test_case.xlsx', "Sheet1", 1, 1)
    # print(res[1].case_title)
    ReadExcel.write_data('./test_case.xlsx', "Sheet1", 1, 3, '注册')
