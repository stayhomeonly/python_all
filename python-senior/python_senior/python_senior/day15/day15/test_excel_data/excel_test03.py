"""
-------------------------------------------------
   Author:Mr.Liu
   date: 2022/8/14-11:33
   contents:
-------------------------------------------------
"""
"""
使用excel对单元格设置字体样式和背景
"""
import openpyxl
from openpyxl.styles import Font, colors, PatternFill

workbook = openpyxl.load_workbook('./test_case.xlsx')  # 加载文件
sheet = workbook["Sheet1"]  # 选择需要操作的sheet页
# 写数据
Color = ['ffffff', '000000']
fille = PatternFill('solid', fgColor=Color[0])  # 设置填充颜色
font = Font(u'微软雅黑', size=20, bold=False, italic=False, strike=False, color=Color[1])  # 设置字体样式

sheet.cell(row=3, column=7, value='失败').fill = fille

sheet.cell(row=3, column=7, value='失败').font = font

workbook.save('./test_case.xlsx')
