"""
---------------------------------
作者: Mr.Liu
创建时间: 2022/8/7-15:10
知识点: 
---------------------------------
"""
import openpyxl  # 用来操作excel

"""
python 操作 excel，使用openpyxl模块
"""
workbook = openpyxl.load_workbook('./test_case.xlsx')  # 加载文件
sheet = workbook["Sheet1"]  # 选择需要操作的sheet页

# 获取cell中的值(excel文件读取的数据除了int/float 类型外，其他全是str字符串类型)
# res = sheet.cell(row=2, column=2).value  # 通过行和列来获取对应的单元格
# print(res, type(res))

# 获取所有的cell
# print(list(sheet.rows))  # 把每一行有值的单元格都装在元组中返回，有几行数据，就有几个元组
res = list(sheet.rows)
print(res[0][1].value)

# 写数据
# res = sheet.cell(row=2, column=3, value='高级班的同学真好看！')
# workbook.save('./test_case.xlsx')
