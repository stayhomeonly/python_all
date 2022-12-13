"""
#@Time：2022/8/22-20:38
#@file：python_basic_56(Excel)
#@Project:python_new_study
#@Content:

"""
from openpyxl import Workbook, load_workbook

# 创建一个文件
# 实例化
# wb=Workbook()
# # 获取当前active的sheet
# sheet=wb.active
#
# print(sheet.title) #Sheet 默认名字
# sheet.title="Fengxin的文件" # 修改文件名字
#
# sheet["B9"]="black girl"
# sheet["C9"]="171,48,83"
# # ⽅式⼆：可以附加⾏，从第⼀列开始附加(从最下⽅空⽩处，最左开始)(可以输⼊多⾏)
# sheet.append(["Rachel",'170','59'])
#
# wb.save("Books2.xlsx") #保存

# 写入数据


# 打开已有的文件
wb = load_workbook("Books2.xlsx")
print(wb.sheetnames)  # ['Fengxin的文件', 'Sheet1']
# print(wb.get_sheet_names())  # ['Fengxin的文件', 'Sheet1'],建议用上面那个
# sheet=wb.get_sheet_by_name("Fengxin的文件")
sheet = wb["Fengxin的文件"]
print(sheet["B5"].value)
print(sheet["B5:B10"])

# for cell in sheet["B5:B10"]:  #获取指定列的数据
#     print(cell[0].value)
print('-' * 50)
# 循环遍历所有打印所有单元格的值
# for row in sheet:
#     # print(row)
#     for cell in row:
#         print(cell.value,end=',')


# 循环遍历指定的行
# for row in sheet.iter_rows(min_row=4,max_row=10,max_col=3):
#     for cell in row:
#         print(cell.value,end=',')
#     print()

# 按列循环
# for column in sheet.columns:
#     for cell in column:
#         print(cell.value, end=',')
#     print()
#指定的列
for column in sheet.iter_cols(min_col=1,max_col=2):
    for cell in column:
        print(cell.value, end=',')
    print()

# 删除工作表
# ⽅式⼀
wb.remove(sheet)
# ⽅式⼆
del wb[sheet]

from openpyxl.styles import Font, colors, Alignment
# 下⾯的代码指定了等线24号，加粗斜体，字体颜⾊红⾊。直接使⽤cell的font属性，将Font对象赋值给
# 它.
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED,
bold=True) # 声明样式
sheet['A1'].font = bold_itatic_24_font # 给单元格设置样式

# 三、对⻬⽅式
# 也是直接使⽤cell的属性aligment，这⾥指定垂直居中和⽔平居中。除了center，还可以使⽤right、left
# 等等参数.
# 设置B1中的数据垂直居中和⽔平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')

# 四、设置⾏⾼&列宽
# 第2⾏⾏⾼
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

